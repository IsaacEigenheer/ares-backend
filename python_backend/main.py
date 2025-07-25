import os
import sys
import numpy as np
import pandas as pd
import fitz
from numba import njit
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
import random
import json
from openpyxl import load_workbook
from shutil import copyfile
from pathlib import Path
import torch
import cv2
from ultralytics import YOLO
import warnings
import openpyxl
import pdfplumber
import re
import tabula
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from google.oauth2.service_account import Credentials
import datetime

os.chdir('/python_backend')

device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

model = YOLO('./Models/newS.pt').to(device)

def calc_size(path):
    with pdfplumber.open(path) as pdf:
        width = pdf.pages[0].width
        height = pdf.pages[0].height
        area = np.sqrt(int(width * height))
        return area, height, width

def start(path, config, current_client, page_):
    global pdf_path
    global filename_id
    page_ = int(page_)
    print('1', flush=True)

    pdf_path = path
    filename_id = (Path(pdf_path)).stem
    area, h_src, w_src = calc_size(path)

    if area < 1500:
        dpi = 400
    elif area >= 1500 and area < 3000:
        dpi = 300
    elif area >= 3000:
        dpi = 175
    
    pdf_document = fitz.open(path)
    if pdf_document.page_count > 0:
        first_page = pdf_document[page_ - 1]
        scale_factor = dpi / 72.0
        image = first_page.get_pixmap(matrix=fitz.Matrix(scale_factor, scale_factor))
        image.save(f'./images/{filename_id}.png')
        image_path = f'./images/{filename_id}.png'
        processar_imagem(image_path, config, current_client, h_src, w_src, scale_factor, pdf_path, page_)
    
    if current_client != 'HPE' and current_client != 'Whirlpool':
        yoloDetect(filename_id)

    if current_client != "Whirlpool" and current_client != "Jacto" and current_client != "CNH":
        excel(path, current_client)

    try:
        make_finalSheet(current_client, filename_id)
    except:
        pass
    
    print('7', flush=True)

    try:
        if os.path.exists('cropped_images'):
            for file in os.listdir('cropped_images'):
                if filename_id in file:
                    os.remove(f'cropped_images/{file}')
        
        if os.path.exists('./crops2'):
            for file in os.listdir('./crops2'):
                if filename_id in file:
                    os.remove(f'./crops2/{file}')

        if os.path.exists('Excel'):
            for file in os.listdir('Excel'):
                if filename_id in file:
                    if 'planilha_final' not in file:
                        os.remove(os.path.join('Excel', file))

        
        if os.path.exists('processed_images'):
            for file in os.listdir('processed_images'):
                if filename_id in file:
                    os.remove(os.path.join('processed_images', file))

        if os.path.exists('images'):
            for file in os.listdir('images'):
                if filename_id in file:
                    os.remove(os.path.join('images', file))
    except:
        pass

    print('8', flush=True)


def processar_imagem(image_path, config, current_client, h_src, w_src, scale_factor, pdf_path, page_):
    image = cv2.imread(image_path)
    h, w, c = image.shape
    detect_lines_and_save(image, os.path.basename(image_path), h, w, config, current_client, h_src, w_src, scale_factor, pdf_path, page_)


def detect_lines_and_save(image, image_name, h, w, config, current_client, h_src, w_src, scale_factor, pdf_path, page_):
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    all_lines = []
    limiares = config['limiares']
    

    for low_trheshold, high_threshold in limiares:
        edges = cv2.Canny(gray, low_trheshold, high_threshold, apertureSize=3)
        lines = cv2.HoughLinesP(edges, 1, np.pi / 180, threshold = 100, minLineLength = config['minLineLenght'], maxLineGap=10)
        if lines is not None:
            all_lines.extend(lines)

    lines_gapMin = []

    lines_vertical_original = []

    lines_horizontal = []

    lines_h_iguais = []

    constante = 0.04314436103201311588575373198723

    for line in all_lines:
        x1, y1, x2, y2 = line[0]
        #cv2.line(image, (x1, y1), (x2, y2), (0,0,255), 6)
        if x1 == x2:
            lines_vertical_original.append(line)
        dist = np.sqrt((x2 - x1)**2 + (y2 - y1)**2)
        dist_horizontal = abs(x2 - x1)
        if dist > (h * constante) and dist_horizontal > (w * 0.01):
            lines_gapMin.append(line)
    
    for line in lines_gapMin:
        x1, y1, x2, y2 = line[0]
        if y1 == y2:
            lines_horizontal.append(line) 

    for line_h in lines_horizontal:
        x1, y1, x2, y2 = line_h[0]
        k = (max(x1, x2) - min(x1, x2))
        lines_h_iguais.append(k)
    
    minimum = 3
    lines_horizontal = np.array(lines_horizontal)
    tolerancia_ = config['tolerancia']
    all_rect = detect_lines_horizontal(lines_horizontal, h, minimum, constante, tolerancia_)

    new_rect = all_rect[((abs(all_rect[:,2] - all_rect[:,0])) > (w*config['minShape'])) & ((abs(all_rect[:,3] - all_rect[:,1])) > (h * config['minShape']))]

    new_rect = new_rect[((abs(new_rect[:,2] - new_rect[:,0])) < (w*config['maxShape'])) | ((abs(new_rect[:,3] - new_rect[:,1])) < (h * config['maxShape']))]

    not_inside = remove_rectangles_inside(new_rect)

    intersectionArea = config['intersectionArea']
    
    non_overlapping_rectangles = remove_overlapping_rectangles(not_inside, intersectionArea)
    non_overlapping_rectangles = np.unique(non_overlapping_rectangles, axis=0)
    t = 0
    for rect in non_overlapping_rectangles:
        x1, y1, x2, y2 = rect

        if h_src < 1500:
            w1 = 0.01
            if current_client == 'Whirlpool':
                w1 = 0.008
        elif h_src >=1500 and h_src < 3500:
            w1 = 0.017
            if current_client == 'Whirlpool':
                w1 = 0.01
        else:
            w1 = 0.025
            if current_client == 'Whirlpool':
                w1 = 0.02
            
        if w_src < 1500:
            w2 = 5
        else:
            w2 = 25

        if (y1 - int(h*w1)) > 0:
            y1 = y1 - int(h*w1)
        
        if(y2 + int(h*w1)) < h:
            y2 = y2 + int(h*w1)

        if (x1 - w2) > 0:
            x1 = x1 - w2
        
        if (x2 + w2) < w:
            x2 = x2 + w2

        if current_client == "Whirlpool" or current_client == "Jacto" or current_client == "CNH":
            output_path = f"./Excel/{t}{image_name}.csv"
            tabula.convert_into(pdf_path, output_path, pages=page_, area=[y1/scale_factor, x1/scale_factor, y2/scale_factor, x2/scale_factor])
            try:
                read_file = pd.read_csv(f'./Excel/{t}{image_name}.csv')
                read_file.to_excel(f'./Excel/{t}{image_name}.xlsx', index=None, header=False)
            except:
                pass

        cropped_image = image[y1:y2, x1:x2]
        output_path = os.path.join("./cropped_images" , f"{t}{image_name}")##################################################################
        cv2.imwrite(output_path, cropped_image)
        t += 1

    output_path = os.path.join(fr"./processed_images/processedimage{filename_id}.png")
    print('2', flush=True)
    cv2.imwrite(output_path, image)

@njit
def detect_lines_horizontal(lines_horizontal, h, minimum, constante, tolerancia_):
    all_rect = []
    for i, line_h1 in enumerate(lines_horizontal):
        x1, y1, x2, y2 = line_h1[0]
        tolerancia = (abs(x2 - x1) * tolerancia_)
        lines_total_x1 = []
        lines_total_y1 = []
        lines_total_x2 = []
        lines_total_y2 = []

        processed_lines = set()
        for j, line_h2 in enumerate(lines_horizontal):
            if j != i and j not in processed_lines:
                x1_, y1_, x2_, y2_ = line_h2[0]
                if (x1 - tolerancia <= x1_ <= x1 + tolerancia) and (x2 - tolerancia <= x2_ <= x2 + tolerancia):
                    lines_total_x1.append(x1_)
                    lines_total_x2.append(x2_)
                    lines_total_y1.append(y1_)
                    lines_total_y2.append(y2_)
                    processed_lines.add(j)
        
        lines_total_y1_sorted = sorted(lines_total_y1)
        lines_total_x1_sorted = [x for _, x in sorted(zip(lines_total_y1, lines_total_x1))]
        lines_total_y2_sorted = [x for _, x in sorted(zip(lines_total_y1, lines_total_y2))]
        lines_total_x2_sorted = [x for _, x in sorted(zip(lines_total_y1, lines_total_x2))]

        lines_total_y1_sorted_restante = sorted(lines_total_y1)
        lines_total_x1_sorted_restante = [x for _, x in sorted(zip(lines_total_y1, lines_total_x1))]
        lines_total_y2_sorted_restante = [x for _, x in sorted(zip(lines_total_y1, lines_total_y2))]
        lines_total_x2_sorted_restante = [x for _, x in sorted(zip(lines_total_y1, lines_total_x2))]

        for i in range(len(lines_total_y1_sorted) -1):
            if abs(lines_total_y1_sorted[i+1] - lines_total_y1_sorted[i] > (h*0.2)):

                lines_total_y1_sorted_restante = lines_total_y1_sorted[i+1:]
                lines_total_x1_sorted_restante = lines_total_x1_sorted[i+1:]
                lines_total_x2_sorted_restante = lines_total_x2_sorted[i+1:]
                lines_total_y2_sorted_restante = lines_total_y2_sorted[i+1:]
                lines_total_y1_sorted = lines_total_y1_sorted[:i]
                lines_total_x1_sorted = lines_total_x1_sorted[:i]
                lines_total_x2_sorted = lines_total_x2_sorted[:i]
                lines_total_y2_sorted = lines_total_y2_sorted[:i]
                break

        if h < 4500:
            minimum = 2
        if len(lines_total_x1_sorted) >= minimum:
            x_min = (min(lines_total_x1_sorted))
            x_max = (max(lines_total_x2_sorted))
            y_min = (min(lines_total_y1_sorted))
            y_max = (max(lines_total_y2_sorted))
            all_rect.append([x_min, y_min, x_max, y_max])
    
        if len(lines_total_x1_sorted_restante) >= minimum:
                x_min = (min(lines_total_x1_sorted_restante))
                x_max = (max(lines_total_x2_sorted_restante))
                y_min = (min(lines_total_y1_sorted_restante))
                y_max = (max(lines_total_y2_sorted_restante))
                all_rect.append([x_min, y_min, x_max, y_max])

    return np.array(all_rect)

@njit
def remove_rectangles_inside(rectangles):
    non_overlapping_rectangles = []
    for rect1 in rectangles:
        for rect2 in rectangles:
            add = True
            if rect2 is not rect1:
                if((rect1[0] > rect2[0]) and (rect1[2] < rect2[2])) and ((rect1[1] > rect2[1]) and (rect1[3] < rect2[3])):
                    add = False
                    break
        if add:        
            non_overlapping_rectangles.append(rect1)

    return non_overlapping_rectangles

@njit
def remove_overlapping_rectangles(rectangles, intersectionArea):
    removed = []
    for rect1 in rectangles:
        keep_rect1 = True
        for rect2 in rectangles:
            if rect2 is not rect1:
                intersection_area = calculate_intersection_area(rect1, rect2)
                area_rect1 = (rect1[2] - rect1[0]) * (rect1[3] - rect1[1])
                area_rect2 = (rect2[2] - rect2[0]) * (rect2[3] - rect2[1])
                if intersection_area > (max(area_rect1, area_rect2) * intersectionArea):
                    if area_rect1 > area_rect2:
                        keep_rect1 = True
                    elif area_rect2 > area_rect1:
                        keep_rect1 = False
                        break
        if keep_rect1:
            removed.append(rect1)
    return removed

@njit
def calculate_intersection_area(rect1, rect2):
    x_overlap = max(0, min(rect1[2], rect2[2]) - max(rect1[0], rect2[0]))
    y_overlap = max(0, min(rect1[3], rect2[3]) - max(rect1[1], rect2[1]))
    return x_overlap * y_overlap

def yoloDetect(filename_id):
    images = []
    for arquivo in os.listdir('cropped_images'):
        if filename_id in arquivo:
            caminho_relativo = os.path.join('cropped_images', arquivo)
            caminho_absoluto = os.path.abspath(caminho_relativo)
            images.append(caminho_absoluto)

    results = model(images, conf=0.7)
    j = 0
    
    for idx, result in enumerate(results):
        image = cv2.imread(images[idx])
        boxes = result.boxes
        for box_idx, box in enumerate(boxes):
            xyxy = box.xyxy[0]
            x1, y1, x2, y2 = xyxy[0].item(), xyxy[1].item(), xyxy[2].item(), xyxy[3].item()
            x1, y1, x2, y2 = int(x1), int(y1), int(x2), int(y2)
            crop_img = image[y1:y2, x1:x2]
            cv2.imwrite(f'./crops2/{j}{filename_id}.png', crop_img)
            j += 1
    print('3', flush=True)

def move_mouse_smoothly(driver, element):
    action = ActionChains(driver)
    action.move_to_element(element)
    action.perform()

def execute_action_with_retry(driver, action_func, max_attempts=2, *args):
    for attempt in range(max_attempts):
        try:
            action_func(driver, *args)
            return  
        except Exception as e:
            time.sleep(random.uniform(2, 5))
    raise



def excel(path, current_client):
    print('4', flush=True)
    caminhos_arquivos = []

    # Lógica de caminhos relativos original mantida
    if current_client != 'HPE':
        path_cropped = './crops2'
    else:
        path_cropped = './cropped_images'

    # Garante que a pasta de origem das imagens exista
    os.makedirs(path_cropped, exist_ok=True)

    for arquivo in os.listdir(path_cropped):
        if filename_id in arquivo:
            caminho_relativo = os.path.join(path_cropped, arquivo)
            caminho_absoluto = os.path.abspath(caminho_relativo)
            caminhos_arquivos.append(caminho_absoluto)

    def convert_to_excel():
        attemps = 0
        while attemps < 3:
            driver = None # Garante que o driver é definido
            try:
                download_directory = os.path.abspath("Excel")
                os.makedirs(download_directory, exist_ok=True)
                
                chrome_options = Options()
                chrome_options.add_argument('--ignore-ssl-errors=yes')
                chrome_options.add_argument('--ignore-certificate-errors')
                chrome_options.add_argument('--no-sandbox')
                chrome_options.add_argument('--headless')
                chrome_options.add_argument('--disable-dev-shm-usage')
                chrome_options.add_argument('--window-size=1920,1080')
                chrome_options.add_extension('./adblock.crx')

                prefs = {
                    "download.default_directory": download_directory,
                    "download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "safeBrowse.enabled": True,
                    "profile.default_content_setting_values.automatic_downloads": 1
                }
                chrome_options.add_experimental_option("prefs", prefs)
                driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)

                driver.get('https://www.google.com/')
                driver.execute_script(f'window.location.href = "https://www.fileeagle.com/pdfeagle/image-to-excel";')
                time.sleep(15)
                file_input = driver.find_element('xpath', '//*[@id="file"]')
                move_mouse_smoothly(driver, file_input)
                qnt_arquivos = len(caminhos_arquivos)
                
                limit = 9 if qnt_arquivos > 9 else qnt_arquivos
                batch = caminhos_arquivos[:limit]
                for arquivo in batch:
                    file_input.send_keys(arquivo)
                
                time.sleep(8)
                driver.execute_script("window.scrollBy(0, 600);")
                language_element = driver.find_element(By.XPATH, '//*[@id="select_language"]/div/div/span/span[1]/span')
                language_element.click()
                english_option = driver.find_element(By.XPATH, '//li[text()="English"]')
                english_option.click()
                time.sleep(4)
                driver.execute_script("window.scrollBy(0, 300);")
                element = driver.find_element(By.ID, 'convert')
                driver.execute_script("arguments[0].click();", element)
                time.sleep(12)
                driver.execute_script("window.scrollBy(0, 100);")
                
                links = []
                u = 0
                while len(links) < limit and u < 120: # Aumentado para 600 como solicitado
                    links = driver.find_elements(By.XPATH, "//a[contains(@title, 'output')]")
                    time.sleep(1)
                    u += 1
                
                links = driver.find_elements(By.XPATH, "//a[contains(@title, 'output')]")
                for index, link in enumerate(links):
                    try:
                        driver.execute_script("""
                        document.querySelectorAll('iframe, .adsbygoogle, [id^="google_ads"]').forEach(ad => ad.remove());
                        """)
                    finally:
                        link.click()
                        time.sleep(9)
                
                time.sleep(6)
                
                if qnt_arquivos > 9:
                    del caminhos_arquivos[0:9]
                    convert_to_excel()
                
                break # Sai do loop de tentativas se foi bem-sucedido

            except Exception as e: # Captura a exceção para logar
                print(f"Erro ao montar driver: {e}")
                attemps += 1
            finally:
                if driver:
                    driver.quit()
    
    convert_to_excel()

def make_finalSheet(current_client, filename_id):
    print('5', flush=True)

    def convert(filename_id):

        SCOPES = ['https://www.googleapis.com/auth/drive']
        CREDENTIALS_FILE = 'API/credentials.json'

        def authenticate():
            creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=SCOPES)
            return creds
        
        def replace_spreadsheet(creds, local_file_path, destination_file_id):
            drive_service = build('drive', 'v3', credentials=creds)
            media = MediaFileUpload(local_file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            drive_service.files().update(
                fileId=destination_file_id,
                media_body=media).execute()
        
        def main():
            try:
                creds = authenticate()
                local_file_path = f'Excel/planilha_final{filename_id}.xlsx'
                destination_file_id = '1JSOfPGPs6Rwq6kij2bZSlALm1DIlwUGeS4cMpWA0rTY'
                replace_spreadsheet(creds, local_file_path, destination_file_id)
            except:
                pass
        caminho_pasta_excel = 'Excel'
        excel_final = 'planilha_final.xlsx'
        excel_modelo = 'cat_importar_dados.xlsx'
        copyfile(excel_modelo, excel_final)
        arquivos_excel = [arquivo for arquivo in os.listdir(caminho_pasta_excel) if arquivo.endswith(".xlsx")]
        wb_destino = load_workbook(excel_final)
        c = 1
        p = 1
        w = 1
        for arquivo in arquivos_excel:
            try:
                if filename_id in arquivo:
                    wb_origem = load_workbook(f'{caminho_pasta_excel}/{arquivo}')
                    table_name = 'tabela'

                    if arquivo != 'planilha_final.xlsx':
                        string1 = 'CIRCUIT DATA TABLE'
                        string2 = 'PARTS LIST'
                        string3 = 'BUNDLE TABLE'

                        df_paracsv = pd.read_excel(os.path.join(caminho_pasta_excel, arquivo))
                        file_path_csv = 'arquivo.csv'
                        df_paracsv.to_csv(file_path_csv, index=False)
                        df_readcsv = pd.read_csv(file_path_csv, header=None, skiprows=1)
                        with open(file_path_csv, 'r', encoding='utf-8') as file:
                            file_content = file.read()
                        
                        if(df_readcsv == string1).any().any():
                            table_name = f'Circuit{c}'
                        elif string1 in file_content:
                            table_name = f'Circuit{c}'
                        
                        if ((df_paracsv == string2)).any().any():
                            table_name = f'Partlist{p}'
                            p += 1
                        elif string2 in file_content:
                            table_name = f'Partlist{p}'
                            p += 1    

                        if(df_readcsv == string3).any().any():
                            table_name = f'Bundle{w}'
                            w += 1
                        elif string3 in file_content:
                                table_name = f'Bundle{w}'
                                w += 1

                        for sheet_name in wb_origem.sheetnames:
                            ws_origem = wb_origem[sheet_name]
                            ws_destino = wb_destino.create_sheet(title=table_name)
                            for row in ws_origem.iter_rows(min_row=1, max_row=1, values_only=True):
                                ws_destino.append(row)

                            for row in ws_origem.iter_rows(min_row=2, values_only=True):
                                ws_destino.append(row)

                    arquivo_excel_path = f'Excel/planilha_final{filename_id}.xlsx'
                    print(f'ExcelFinal {arquivo_excel_path}', flush=True)
                    wb_destino.save(arquivo_excel_path)
                    df_final = pd.read_excel(arquivo_excel_path, sheet_name='DADOS')    
                    df_final.at[3, 'X'] = c
                    df_final.at[18, 'X'] = p
                    print('6', flush=True)
            except:
                continue
        main()   


    def generic_convert(filename_id):
        caminho_pasta_excel = 'Excel'
        excel_final = 'generic_spreadsheet.xlsx'
        arquivos_excel = [arquivo for arquivo in os.listdir(caminho_pasta_excel) if arquivo.endswith(".xlsx")]
        wb_destino = load_workbook(excel_final)
        for arquivo in arquivos_excel:
            if filename_id in arquivo:
                wb_origem = load_workbook(f'{caminho_pasta_excel}/{arquivo}')
                table_name = 'tabela'

                if arquivo != 'planilha_final.xlsx':

                    df_paracsv = pd.read_excel(os.path.join(caminho_pasta_excel, arquivo))
                    file_path_csv = 'arquivo.csv'
                    df_paracsv.to_csv(file_path_csv, index=False)

                    for sheet_name in wb_origem.sheetnames:
                        ws_origem = wb_origem[sheet_name]
                        ws_destino = wb_destino.create_sheet(title=table_name)
                        for row in ws_origem.iter_rows(min_row=1, max_row=1, values_only=True):
                            ws_destino.append(row)

                        for row in ws_origem.iter_rows(min_row=2, values_only=True):
                            ws_destino.append(row)

                arquivo_excel_path = f'Excel/planilha_final{filename_id}.xlsx'
                print(f'ExcelFinal {arquivo_excel_path}', flush=True)
                wb_destino.save(arquivo_excel_path)
                print('6', flush=True)
            
    def whirlpoolConvert(filename_id):
        caminho_pasta_excel = 'Excel'
        arquivos_excel = [arquivo for arquivo in os.listdir(caminho_pasta_excel) if arquivo.endswith(".xlsx")]

        # Criar workbook final com aba de resumo
        wb_destino = openpyxl.Workbook()
        ws_resumo = wb_destino.active
        ws_resumo.title = "Resumo_Cores"
        ws_resumo.append(["Cor", "UL STYLE", "WIRE GAUGE", "TEMP RATING", "Código"])  # Cabeçalho da aba consolidada

        # Dicionário de conversão de cores
        color_map = {
            'YELLOW': 'YL', 'BLUE': 'BU', 'GRAY': 'GY', 'GREY': 'GY',
            'ORANGE': 'OR', 'BROWN': 'BR', 'BLACK': 'BK', 'GREEN': 'GN', 'RED': 'RD',
            'PINK': 'PK', 'GREEN/YELLOW': 'GN/YL', 'VIOLET': 'VT', 'LT BLUE': 'LT BU', 'LTBLUE': 'LT BU', 'WHITE': 'WH', 'PURPLE': 'PU',
            'YELLOW/BLACK': 'YL/BK', 'GREY/WHITE': 'GY/WH', 'GRAY/WHITE': 'GY/WH', 'BROWN/BLACK': 'BR/BK', 'GREY/BLACK': 'GY/BK', 'GRAY/BLACK': 'GY/BK',
            'ORANGE/BLACK': 'OR/BK', 'BLUE/BLACK': 'BU/BK', 'BLUE/WHITE': 'BU/WH'
        }

        for arquivo in arquivos_excel:
            # Ignorar a planilha final para não processá-la novamente
            if filename_id in arquivo and arquivo != f'planilha_final{filename_id}.xlsx':
                caminho_arquivo = os.path.join(caminho_pasta_excel, arquivo)
                wb_origem = load_workbook(caminho_arquivo, data_only=True)

                for aba_nome in wb_origem.sheetnames:
                    # Ignorar a aba "Resumo_Cores" se ela existir na planilha origem
                    if aba_nome == "Resumo_Cores":
                        continue

                    ws_origem = wb_origem[aba_nome]
                    # Copiar conteúdo da aba original para o workbook final (com limite de 31 caracteres no nome)
                    ws_dest = wb_destino.create_sheet(title=aba_nome[:31])
                    for row in ws_origem.iter_rows(values_only=True):
                        ws_dest.append(row)

                    # Identificar a linha de cabeçalho e as colunas relevantes
                    header_found = False
                    header_index = None
                    colunas_encontradas = {}
                    for i, row in enumerate(ws_origem.iter_rows(values_only=True)):
                        if row:
                            for j, cell in enumerate(row):
                                if isinstance(cell, str):
                                    cell_norm = cell.strip().lower().replace(" ", "")
                                    if cell_norm in ["wirecolor", "color"]:
                                        colunas_encontradas["cor"] = j
                                    elif cell_norm in ["ulstyle", "wirestyle"]:
                                        colunas_encontradas["ul"] = j
                                    elif cell_norm in ["wiregauge", "awg"]:
                                        colunas_encontradas["wire"] = j
                                    elif cell_norm in ["temp", "temprating"]:
                                        colunas_encontradas["temp"] = j
                            # Verifica se encontrou todas as colunas necessárias
                            if all(key in colunas_encontradas for key in ["cor", "ul", "wire", "temp"]):
                                header_found = True
                                header_index = i
                                break

                    # Se o cabeçalho for identificado, processa as linhas seguintes
                    if header_found and header_index is not None:
                        rows = list(ws_origem.iter_rows(values_only=True))
                        for row in rows[header_index+1:]:
                            if not any(row):
                                continue  # Pular linhas vazias

                            # Obter valores de forma segura, verificando se o índice existe na linha
                            cor_val = row[colunas_encontradas["cor"]] if len(row) > colunas_encontradas["cor"] else ""
                            cor_original = str(cor_val).strip() if cor_val is not None else ""
                            cor_convertida = color_map.get(cor_original.upper(), cor_original)

                            ul_val = row[colunas_encontradas["ul"]] if len(row) > colunas_encontradas["ul"] else ""
                            ul_style = str(ul_val).strip() if ul_val is not None else ""

                            wire_val = row[colunas_encontradas["wire"]] if len(row) > colunas_encontradas["wire"] else ""
                            wire_gauge = str(wire_val).strip() if wire_val is not None else ""

                            temp_val = row[colunas_encontradas["temp"]] if len(row) > colunas_encontradas["temp"] else ""
                            temp_rating = str(temp_val).strip() if temp_val is not None else ""

                            # Criar código no formato: wiregauge_cor_ulstyle
                            codigo = f"{wire_gauge}_{cor_convertida}_{ul_style}"

                            ws_resumo.append([cor_convertida, ul_style, wire_gauge, temp_rating, codigo])

        # Salvar a planilha final contendo todas as abas
        arquivo_excel_path = f'Excel/planilha_final{filename_id}.xlsx'
        print(f'ExcelFinal {arquivo_excel_path}', flush=True)
        wb_destino.save(arquivo_excel_path)
        print('6', flush=True)


    warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.styles.stylesheet")
    if current_client == 'Caterpillar':
        convert(filename_id)
    elif current_client == 'Whirlpool':
        whirlpoolConvert(filename_id)
    else:
        generic_convert(filename_id)


    

if __name__ == '__main__':
    arg1 = sys.argv[1]
    arg2 = sys.argv[2]
    with open('./config.json') as json_data:
        data = json.load(json_data,)
        config = data['customers'][arg2]
    arg3 = sys.argv[3]
    start(arg1, config, arg2, arg3)

