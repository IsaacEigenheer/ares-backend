import os
import fitz
import cv2
import numpy as np
from numba import njit
from pathlib import Path
from selenium.webdriver.common.action_chains import ActionChains
import pdfplumber
import random
import time
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from shutil import copyfile
import pandas as pd
import re
import openpyxl
from typing import Tuple

class Detect_Image():
    def __init__(self, path: str, config: dict, current_client: str, page: str):
        self.pdf_path = path
        self.config = config
        self.current_client = current_client
        self.page = int(page)
        self.filename_id = (Path(self.pdf_path)).stem


    def calcPdfSize(self) -> Tuple[float, float, float]:
        with pdfplumber.open(self.pdf_path) as pdf:
            width_source = pdf.pages[0].width
            height_source = pdf.pages[0].height
            area = np.sqrt(int(width_source * height_source))
            return area, height_source, width_source

    def calcDpi(self) -> int:
        if self.area < 1500:
            return 400
        elif self.area >= 1500 and self.area < 3000:
            return 300
        elif self.area >= 3000:
            return 175
    
    def calcParameters(self) -> str:
        pdf_document = fitz.open(self.pdf_path)
        if pdf_document.page_count > 0:
            first_page = pdf_document[self.page - 1]
            scale_factor = self.dpi / 72
            pdf_image = first_page.get_pixmap(matrix=fitz.Matrix(scale_factor, scale_factor))
            pdf_image.save(f'./images/{self.filename_id}.png')
            pdf_image_path = f'./images/{self.filename_id}.png'
            return pdf_image_path
    
    def processarImagem(self) -> Tuple[np.array, float, float, str]:
        image_src = cv2.imread(self.pdf_image_path)
        height_img, width_img, _ = image_src.shape
        image_name = os.path.basename(self.pdf_image_path)
        return image_src, height_img, width_img, image_name
    
    def detect_lines_and_save(self) -> Tuple[np.array, int, float, float]:
        gray = cv2.cvtColor(self.image_src, cv2.COLOR_BGR2GRAY)
        all_lines = []
        limiares = self.config['limiares']

        for low_trheshold, high_threshold in limiares:
            edges = cv2.Canny(gray, low_trheshold, high_threshold, apertureSize=3)
            lines = cv2.HoughLinesP(edges, 1, np.pi / 180, threshold = 100, minLineLength = self.config['minLineLenght'], maxLineGap=10)
            if lines is not None:
                all_lines.extend(lines)
            
        lines_gapMin = []

        lines_vertical_original = []

        lines_horizontal = []

        lines_h_iguais = []

        constante = 0.04314436103201311588575373198723

        for line in all_lines:
            x1, y1, x2, y2 = line[0]
            if x1 == x2:
                lines_vertical_original.append(line)
            dist = np.sqrt((x2 - x1)**2 + (y2 - y1)**2)
            dist_horizontal = abs(x2 - x1)
            if dist > (self.height_img * constante) and dist_horizontal > (self.width_img * 0.01):
                lines_gapMin.append(line)
            
        for line in lines_gapMin:
            x1, y1, x2, y2 = line[0]
            if y1 == y2:
                lines_horizontal.append(line)
        
        for line_h in lines_horizontal:
            x1, y1, x2, y2 = line_h[0]
            k = (max(x1, x2) - min(x1, x2))
            lines_h_iguais.append(k)
        
        minimum = 3
        lines_horizontal = np.array(lines_horizontal)
        tolerancia_ = self.config['tolerancia']

        return lines_horizontal, minimum, constante, tolerancia_

    @staticmethod
    @njit
    def calcAll_rect(lines_horizontal: np.array, tolerancia_: float, height_img: float, minimum: int) -> np.array:
        all_rect = []
        for i, line_h1 in enumerate(lines_horizontal):
            x1, y1, x2, y2 = line_h1[0]
            tolerancia = (abs(x2 - x1) * tolerancia_)
            lines_total_x1 = []
            lines_total_y1 = []
            lines_total_x2 = []
            lines_total_y2 = []

            processed_lines = set()
            for j, line_h2 in enumerate(lines_horizontal):
                if j != i and j not in processed_lines:
                    x1_, y1_, x2_, y2_ = line_h2[0]
                    if (x1 - tolerancia <= x1_ <= x1 + tolerancia) and (x2 - tolerancia <= x2_ <= x2 + tolerancia):
                        lines_total_x1.append(x1_)
                        lines_total_x2.append(x2_)
                        lines_total_y1.append(y1_)
                        lines_total_y2.append(y2_)
                        processed_lines.add(j)
            
            lines_total_y1_sorted = sorted(lines_total_y1)
            lines_total_x1_sorted = [x for _, x in sorted(zip(lines_total_y1, lines_total_x1))]
            lines_total_y2_sorted = [x for _, x in sorted(zip(lines_total_y1, lines_total_y2))]
            lines_total_x2_sorted = [x for _, x in sorted(zip(lines_total_y1, lines_total_x2))]

            lines_total_y1_sorted_restante = sorted(lines_total_y1)
            lines_total_x1_sorted_restante = [x for _, x in sorted(zip(lines_total_y1, lines_total_x1))]
            lines_total_y2_sorted_restante = [x for _, x in sorted(zip(lines_total_y1, lines_total_y2))]
            lines_total_x2_sorted_restante = [x for _, x in sorted(zip(lines_total_y1, lines_total_x2))]

            for i in range(len(lines_total_y1_sorted) -1):
                if abs(lines_total_y1_sorted[i+1] - lines_total_y1_sorted[i] > (height_img *0.2)):

                    lines_total_y1_sorted_restante = lines_total_y1_sorted[i+1:]
                    lines_total_x1_sorted_restante = lines_total_x1_sorted[i+1:]
                    lines_total_x2_sorted_restante = lines_total_x2_sorted[i+1:]
                    lines_total_y2_sorted_restante = lines_total_y2_sorted[i+1:]
                    lines_total_y1_sorted = lines_total_y1_sorted[:i]
                    lines_total_x1_sorted = lines_total_x1_sorted[:i]
                    lines_total_x2_sorted = lines_total_x2_sorted[:i]
                    lines_total_y2_sorted = lines_total_y2_sorted[:i]
                    break
            
            if height_img < 4500:
                minimum = 2

            if len(lines_total_x1_sorted) >= minimum:

                x_min = (min(lines_total_x1_sorted))
                x_max = (max(lines_total_x2_sorted))
                y_min = (min(lines_total_y1_sorted))
                y_max = (max(lines_total_y2_sorted))
                all_rect.append([x_min, y_min, x_max, y_max])
            
            if len(lines_total_x1_sorted_restante) >= minimum:
                x_min = (min(lines_total_x1_sorted_restante))
                x_max = (max(lines_total_x2_sorted_restante))
                y_min = (min(lines_total_y1_sorted_restante))
                y_max = (max(lines_total_y2_sorted_restante))
                all_rect.append([x_min, y_min, x_max, y_max])
        
        return np.array(all_rect)
    
    @staticmethod
    @njit
    def processamento_retangulos(new_rect: np.array, intersectionArea: float) -> np.array:
        non_overlapping_rectangles = []
        for rect1 in new_rect:
            for rect2 in new_rect:
                add = True
                if rect2 is not rect1:
                    if((rect1[0] > rect2[0]) and (rect1[2] < rect2[2])) and ((rect1[1] > rect2[1]) and (rect1[3] < rect2[3])):
                        add = False
                        break
            if add:
                non_overlapping_rectangles.append(rect1)
        
        not_inside = non_overlapping_rectangles

        removed = []
        for rect1 in not_inside:
            keep_rect1 = True
            for rect2 in not_inside:
                if rect2 is not rect1:
                    x_overlap = max(0, min(rect1[2], rect2[2]) - max(rect1[0], rect2[0]))
                    y_overlap = max(0, min(rect1[3], rect2[3]) - max(rect1[1], rect2[1]))
                    intersection_area = (x_overlap * y_overlap)
                    area_rect1 = (rect1[2] - rect1[0]) * (rect1[3] - rect1[1])
                    area_rect2 = (rect2[2] - rect2[0]) * (rect2[3] - rect2[1])
                    if intersection_area > (max(area_rect1, area_rect2) * intersectionArea):
                        if area_rect1 > area_rect2:
                            keep_rect1 = True
                        elif area_rect2 > area_rect1:
                            keep_rect1 = False
                            break
            if keep_rect1:
                removed.append(rect1)
        
        non_overlapping_rectangles = removed
        return non_overlapping_rectangles
    
    def save_detections(self):
        t = 0
        for rect in self.non_overlapping_rectangles:
            x1, y1, x2, y2 = rect
            if self.height_source < 1500:
                w1 = 0.01
            elif self.height_source >= 1500 and self.height_source < 3500:
                w1 = 0.017
            else:
                w1 = 0.025

            if self.width_source < 1500:
                w2 = 5
            else:
                w2 = 25
            
            if (y1 - int(self.height_img * w1)) > 0:
                y1 = y1 - int(self.height_img * w1)
            
            if(y2 + int(self.height_img * w1)) < self.height_img:
                y2 = y2 + int(self.height_img * w1)

            if (x1 - w2) > 0:
                x1 = x1 - w2
            
            if (x2 + w2) < self.width_img:
                x2 = x2 + w2
        
            cropped_image = self.image_src[y1:y2, x1:x2]
            output_path = os.path.join("./cropped_images" , f"{t}{self.image_name}")##################################################################
            cv2.imwrite(output_path, cropped_image)
            t += 1
       
        output_path = os.path.join(fr"./processed_images/processedimage{self.filename_id}.png")
        print('2', flush=True)
        cv2.imwrite(output_path, self.image_src)

class Convert_Excel():
    def __init__(self, current_client: str, filename_id: str):
        print('4', flush=True)
        self.caminhos_arquivos = []
        if current_client != 'HPE' and current_client != 'Whirlpool':
            self.path_cropped = './crops2'
        else:
            self.path_cropped = './cropped_images'
        
        for arquivo in os.listdir(self.path_cropped):
            if filename_id in arquivo:
                caminho_relativo = os.path.join(self.path_cropped, arquivo)
                caminho_absoluto = os.path.abspath(caminho_relativo)
                self.caminhos_arquivos.append(caminho_absoluto)

    def move_mouse_smoothly(self, driver, element):
        action = ActionChains(driver)
        action.move_to_element(element)
        action.perform()
    
    def execute_action_with_retry(self, driver, action_func, max_attempts=2, *args):
        for attempt in range(max_attempts):
            try:
                action_func(driver, *args)
                return  
            except Exception as e:
                time.sleep(random.uniform(2, 5))
        raise
    
    def excel(self):
        attemps = 0
        while attemps < 3:
            try:
                download_directory = os.path.abspath("Excel")
                
                chrome_options = Options()
                chrome_options.add_argument('--ignore-ssl-errors=yes')
                chrome_options.add_argument('--ignore-certificate-errors')
                chrome_options.add_argument('--no-sandbox')
                chrome_options.add_argument('--headless')
                chrome_options.add_argument('--disable-dev-shm-usage')
                chrome_options.add_argument('--window-size=1920,1080')

                prefs = {
                    "download.default_directory": download_directory,
                    "download.prompt_for_download": False,
                    "download.directory_upgrade": True,
                    "safebrowsing.enabled": True,
                    "profile.default_content_setting_values.automatic_downloads": 1
                }

                chrome_options.add_experimental_option("prefs", prefs)
                driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chrome_options)

                driver.get('https://www.google.com/')
                driver.execute_script(f'window.location.href = "https://www.fileeagle.com/pdfeagle/image-to-excel";')
                time.sleep(15)
                file_input = driver.find_element('xpath', '//*[@id="file"]')
                self.move_mouse_smoothly(driver, file_input)
                qnt_arquivos = len(self.caminhos_arquivos)
                if qnt_arquivos <= 11:
                    for arquivo in self.caminhos_arquivos:
                        file_input.send_keys(arquivo)
                fileLoop = 0
                if qnt_arquivos > 11:
                    for arquivo in self.caminhos_arquivos:
                        if fileLoop < 11:
                                file_input.send_keys(arquivo)
                                
                                fileLoop += 1 
                time.sleep(8)
                driver.execute_script("window.scrollBy(0, 600);")
                language_element = driver.find_element(By.XPATH, '//*[@id="select_language"]/div/div/span/span[1]/span')
                language_element.click()
                english_option = driver.find_element(By.XPATH, '//li[text()="English"]')
                english_option.click()
                time.sleep(4)
                driver.execute_script("window.scrollBy(0, 300);")
                element = driver.find_element(By.ID, 'convert')
                driver.execute_script("arguments[0].click();", element)
                time.sleep(12)
                driver.execute_script("window.scrollBy(0, 100);")
                links = driver.find_elements(By.XPATH, "//a[contains(@title, 'output')]")
                if qnt_arquivos <= 11:
                    u = 0
                    while len(links) < (qnt_arquivos) and u < 30800:
                        links = driver.find_elements(By.XPATH, "//a[contains(@title, 'output')]")
                        u +=1
                if qnt_arquivos > 11:
                    u = 0
                    while len(links) < 11 and u < 30800:
                        links = driver.find_elements(By.XPATH, "//a[contains(@title, 'output')]")
                        u +=1     
                links = driver.find_elements(By.XPATH, "//a[contains(@title, 'output')]")
                for index, link in enumerate(links):
                    link.click()
                    time.sleep(3)
                time.sleep(6)
                driver.quit()
                time.sleep(3)
                if qnt_arquivos > 11:
                    del self.caminhos_arquivos[0:11]
                    self.excel()
                break
            except:
                attemps += 1

class Final_Sheet():
    print('5', flush=True)
    def __init__(self, current_client: str, filename_id: str):
        self.current_client = current_client
        self.filename_id = filename_id
    
    def cat_convert(self):
        caminho_pasta_excel = './Excel'
        excel_final = './planilha_final.xlsx'
        excel_modelo = './cat_importar_dados.xlsx'
        copyfile(excel_modelo, excel_final)
        arquivos_excel = [arquivo for arquivo in os.listdir(caminho_pasta_excel) if arquivo.endswith(".xlsx")]
        wb_destino = load_workbook(excel_final)
        c = 1
        p = 1
        w = 1
        for arquivo in arquivos_excel:
            if self.filename_id in arquivo:
                wb_origem = load_workbook(f'{caminho_pasta_excel}/{arquivo}')
                table_name = 'tabela'
                
                if arquivo != 'planilha_final.xlsx':
                    string1 = 'CIRCUIT DATA TABLE'
                    string2 = 'PARTS LIST'
                    string3 = 'BUNDLE TABLE'

                    df_paracsv = pd.read_excel(os.path.join(caminho_pasta_excel, arquivo))
                    file_path_csv = 'arquivo.csv'
                    df_paracsv.to_csv(file_path_csv, index=False)
                    df_readcsv = pd.read_csv(file_path_csv, header=None, skiprows=1)
                    with open(file_path_csv, 'r', encoding='utf-8') as file:
                        file_content = file.read()

                    if(df_readcsv == string1).any().any():
                        table_name = f'Circuit{c}'
                    
                    elif string1 in file_content:
                        table_name = f'Circuit{c}'
                    
                    if ((df_paracsv == string2)).any().any():
                        table_name = f'Partlist{p}'
                        p += 1
                    elif string2 in file_content:
                        table_name = f'Partlist{p}'
                        p += 1  
                
                    if(df_readcsv == string3).any().any():
                        table_name = f'Bundle{w}'
                        w += 1
                    elif string3 in file_content:
                            table_name = f'Bundle{w}'
                            w += 1

                    for sheet_name in wb_origem.sheetnames:
                        ws_origem = wb_origem[sheet_name]
                        ws_destino = wb_destino.create_sheet(title=table_name)
                        for row in ws_origem.iter_rows(min_row=1, max_row=1, values_only=True):
                            ws_destino.append(row)
                        for row in ws_origem.iter_rows(min_row=2, values_only=True):
                            ws_destino.append(row)

                arquivo_excel_path = f'./Excel/planilha_final{self.filename_id}.xlsx'
                #print(f'./ExcelFinal {arquivo_excel_path}', flush=True)
                wb_destino.save(arquivo_excel_path)
                df_final = pd.read_excel(arquivo_excel_path, sheet_name='DADOS')    
                df_final.at[3, 'X'] = c
                df_final.at[18, 'X'] = p
                print('6', flush=True)
    

    def generic_convert(self):
        caminho_pasta_excel = 'Excel'
        excel_final = 'generic_spreadsheet.xlsx'
        arquivos_excel = [arquivo for arquivo in os.listdir(caminho_pasta_excel) if arquivo.endswith(".xlsx")]
        wb_destino = load_workbook(excel_final)
        for arquivo in arquivos_excel:
            if self.filename_id in arquivo:
                wb_origem = load_workbook(f'{caminho_pasta_excel}/{arquivo}')
                table_name = 'tabela'

                if arquivo != 'planilha_final.xlsx':
                    df_paracsv = pd.read_excel(os.path.join(caminho_pasta_excel, arquivo))
                    file_path_csv = './arquivo.csv'
                    df_paracsv.to_csv(file_path_csv, index=False)
                
                    for sheet_name in wb_origem.sheetnames:
                        ws_origem = wb_origem[sheet_name]
                        ws_destino = wb_destino.create_sheet(title=table_name)
                        for row in ws_origem.iter_rows(min_row=1, max_row=1, values_only=True):
                            ws_destino.append(row)

                        for row in ws_origem.iter_rows(min_row=2, values_only=True):
                            ws_destino.append(row)
                
                arquivo_excel_path = f'./Excel/planilha_final{self.filename_id}.xlsx'
                wb_destino.save(arquivo_excel_path)
                print('6', flush=True)
    

    def whirlpool_convert(self):
        caminho_pasta_excel = './Excel'
        excel_final = './wp_spreadsheet.xlsx'
        arquivos_excel = [arquivo for arquivo in os.listdir(caminho_pasta_excel) if arquivo.endswith(".xlsx")]
        wb_destino = load_workbook(excel_final)

        for arquivo in arquivos_excel:
            proc_data = [[], [], [], []]
            if self.filename_id in arquivo:
                wb_origem = load_workbook(f'{caminho_pasta_excel}/{arquivo}')
                table_name = 'tabela'

                if arquivo != 'planilha_final.xlsx':
                    df_paracsv = pd.read_excel(os.path.join(caminho_pasta_excel, arquivo))

                    #WHIRLPOOL PART NUMBER VERIFICAÇÃO
                    column_name = None
                    normalized_columns = {}
                    for index in range(len(df_paracsv)):
                        for col in df_paracsv.columns:
                            normalized_value = str(df_paracsv.at[index, col]).lower().replace(' ', '').replace('i', '1').replace('l', '1')
                            if 'wh1r1poo1partnumber' == normalized_value:
                                column_name = col 
                            if column_name and column_name in df_paracsv.columns:
                                df_paracsv[column_name] = df_paracsv[column_name].astype(str).str.replace(' ', '').str.replace('i', '1', case=False).str.replace('l', '1', case=False)
                        
                    for col in df_paracsv.columns:
                        normalized_col = str(col).lower().replace(' ', '').replace('i', '1').replace('l', '1')
                        normalized_columns[col] = normalized_col

                        if normalized_col == 'wh1r1poo1partnumber':
                            column_name = col 
                        
                        if column_name:
                            df_paracsv[column_name] = df_paracsv[column_name].astype(str).str.replace(' ', '').str.replace('i', '1', case=False).str.replace('l', '1', case=False)

                    #MANUFACTURER PART NUMBER VERIFICAÇÃO
                    column_name = None
                    normalized_columns = {}
                    for index in range(len(df_paracsv)):
                        for col in df_paracsv.columns:
                            normalized_value = str(df_paracsv.at[index, col]).lower().replace(' ', '').replace('i', '1').replace('l', '1')

                            if 'manufacturerpartnumber' == normalized_value:
                                column_name = col 
                    
                            if column_name and column_name in df_paracsv.columns:
                                df_paracsv[column_name] = df_paracsv[column_name].astype(str).str.replace(' ', '').str.replace('i', '1', case=False).str.replace('l', '1', case=False)

                    for col in df_paracsv.columns:
                        normalized_col = str(col).lower().replace(' ', '').replace('i', '1').replace('l', '1')
                        normalized_columns[col] = normalized_col
                        
                        if normalized_col == 'manufacturerpartnumber':
                            column_name = col 

                        if column_name:
                            df_paracsv[column_name] = df_paracsv[column_name].astype(str).str.replace(' ', '').str.replace('i', '1', case=False).str.replace('l', '1', case=False)
                    
                    #TEMPERATURA VERIFICAÇÃO
                    normalized_columns = {}
                    column_name = None

                    for col in df_paracsv.columns:
                        normalized_col = str(col).lower().replace(' ', '').replace('i', '1').replace('l', '1')
                        normalized_columns[col] = normalized_col
                        
                        if 'c1ass' in normalized_col:
                            column_name = col 

                        if column_name:
                            df_paracsv[column_name] = df_paracsv[column_name].astype(str).str.replace(' ', '').str.replace(r'\.\d+', '', regex=True)
                    if column_name:
                        for item in df_paracsv[column_name]:
                            proc_data[3].append(item)
                    
                    #COR DO CABO VERIFICAÇÃO
                    normalized_columns = {}
                    column_name = None

                    for col in df_paracsv.columns:
                        normalized_col = str(col).lower().replace(' ', '').replace('i', '1').replace('l', '1')
                        normalized_columns[col] = normalized_col
                        
                        if normalized_col == 'w1reco1or' or normalized_col == 'co1or':
                            column_name = col 

                        if column_name:
                            df_paracsv[column_name] = df_paracsv[column_name].astype(str).str.replace(' ', '')
                    if column_name:
                        for item in df_paracsv[column_name]:
                            if item == 'YELLOW':
                                item = 'AM'
                            elif item == 'BLUE':
                                item = 'AZ'
                            elif item == 'WHITE':
                                item = 'BR'
                            elif item == 'GRAY' or item == 'GREY':
                                item = 'CZ'
                            elif item == 'ORANGE':
                                item = 'LJ'
                            elif item == 'BROWN':
                                item = 'MR'
                            elif item == 'BLACK':
                                item = 'PR'
                            elif item == 'GREEN':
                                item = 'VD'
                            elif item == 'RED':
                                item = 'VM'
                            elif item == 'PINK':
                                item = 'RO'
                            elif item == 'GREEN/YELLOW':
                                item = 'VD/AM'
                            elif item == 'VIOLET':
                                item = 'VI'
                            elif item == 'LT BLUE' or item == 'LTBLUE':
                                item = 'AZ CL'

                            proc_data[0].append(item)
                    
                    #BITOLA VERIFICAÇÃO
                    normalized_columns = {}
                    column_name = None

                    for col in df_paracsv.columns:
                        normalized_col = str(col).lower().replace(' ', '').replace('i', '1').replace('l', '1')
                        normalized_columns[col] = normalized_col
                        
                        if normalized_col == 'w1regauge':
                            column_name = col 

                        if column_name:
                            df_paracsv[column_name] = df_paracsv[column_name].astype(str).str.replace(' ', '').str.replace('i', '1', case=False).str.replace('l', '1', case=False)
                    if column_name:
                        def format_value(item):
                            match = re.match(r'^\s*(\d+)([.,]\d+)?\s*[a-zA-Z]*', item)
                            if match:
                                integer_part = match.group(1)
                                decimal_part = match.group(2)

                                if decimal_part:
                                    decimal_part = decimal_part.replace('.', ',')
                                    decimal_part = decimal_part[:3]
                                    if len(decimal_part) < 3:
                                        decimal_part = decimal_part + '0'
                                else:
                                    decimal_part = ',00'

                                formatted_value = f"{integer_part}{decimal_part} mm²"
                                return formatted_value.strip()
                            return None
                        for item in df_paracsv[column_name]:
                            formatted_item = format_value(item)
                            proc_data[1].append(formatted_item)

                    if (len(proc_data[0]) or len(proc_data[1]) == 0):
                        def format_value(item):
                            match = re.match(r'^\s*(\d+)([.,]\d+)?\s*[a-zA-Z]*', item)
                            if match:
                                integer_part = match.group(1)
                                decimal_part = match.group(2)

                                if decimal_part:
                                    decimal_part = decimal_part.replace('.', ',')
                                    decimal_part = decimal_part[:3]
                                    if len(decimal_part) < 3:
                                        decimal_part = decimal_part + '0'
                                else:
                                    decimal_part = ',00'

                                formatted_value = f"{integer_part}{decimal_part} mm²"
                                return formatted_value.strip()
                            return None

                        normalized_columns = {}
                        spool_column_name = None

                        for col in df_paracsv.columns:
                            normalized_col = str(col).lower().replace(' ', '').replace('i', '1').replace('l', '1')
                            normalized_columns[col] = normalized_col
                            print(f'Normalized_col: {normalized_col}')
                            if 'spoo1name' in normalized_col:
                                spool_column_name = col

                            if spool_column_name and spool_column_name in df_paracsv.columns:
                                        df_paracsv[spool_column_name] = df_paracsv[spool_column_name].astype(str).str.replace(' ', '')

                        if spool_column_name == None:
                            for index in range(len(df_paracsv)):
                                for col in df_paracsv.columns:
                                    normalized_value = str(df_paracsv.at[index, col]).lower().replace(' ', '').replace('i', '1').replace('l', '1')
                                    if 'spoo1name' in normalized_value:
                                        spool_column_name = col
                                        
                            
                                    if spool_column_name and spool_column_name in df_paracsv.columns:
                                        df_paracsv[spool_column_name] = df_paracsv[spool_column_name].astype(str).str.replace(' ', '')

                        # Se a coluna foi encontrada
                        if spool_column_name:
                            print('Valor spool name encontrado')
                            # Normalizando os dados na coluna
                            df_paracsv[spool_column_name] = df_paracsv[spool_column_name].astype(str).str.replace(' ', '')

                            for item in df_paracsv[spool_column_name]:
                                # Usando regex para extrair bitola e cor
                                match = re.search(r'(\d+(?:\.\d+)?)(?:\s*/\s*|\s+|)([A-Z]+)(?:\s+|/)(\d+)', item)
                                print(f'Item: {item}')
                                if match:
                                    bitola = match.group(1)  # Primeiro grupo: a bitola
                                    cor = match.group(2)      # Segundo grupo: a cor
                                    temperatura = match.group(3)  # Terceiro grupo: a temperatura

                                    # Processando a cor
                                    cor_mapeada = {
                                        'YELLOW': 'AM',
                                        'BLUE': 'AZ',
                                        'WHITE': 'BR',
                                        'GRAY': 'CZ',
                                        'GREY': 'CZ',
                                        'ORANGE': 'LJ',
                                        'BROWN': 'MR',
                                        'BLACK': 'PR',
                                        'GREEN': 'VD',
                                        'RED': 'VM',
                                        'PINK': 'RO',
                                        'GREEN/YELLOW': 'VD/AM',
                                        'VIOLET': 'VI',
                                        'LT BLUE': 'AZ CL',
                                        'LTBLUE' : 'AZ CL'
                                    }

                                    cor_final = cor_mapeada.get(cor, None)

                                    # Formatando a bitola
                                    formatted_bitola = format_value(bitola)

                                    # Adicionando os valores processados
                                    if cor_final:
                                        proc_data[0].append(cor_final)
                                    if formatted_bitola:
                                        proc_data[1].append(formatted_bitola)
                                    if temperatura:
                                        proc_data[3].append(temperatura)  # Adicionando a temperatura

                    #CONECTOR VERIFICAÇÃO
                    column_name = None
                    normalized_columns = {}
                    for index in range(len(df_paracsv)):
                        for col in df_paracsv.columns:
                            normalized_value = str(df_paracsv.at[index, col]).lower().replace(' ', '').replace('i', '1').replace('l', '1')

                            if 'connector' == normalized_value:
                                column_name = col 
                                    
                            if column_name and column_name in df_paracsv.columns:
                                df_paracsv[column_name] = df_paracsv[column_name].astype(str).str.replace(' ', '').str.replace('i', '1', case=False).str.replace('l', '1', case=False)

                    for col in df_paracsv.columns:
                        normalized_col = str(col).lower().replace(' ', '').replace('i', '1').replace('l', '1')
                        normalized_columns[col] = normalized_col
                        
                        if normalized_col == 'connector':
                            column_name = col 

                        if column_name:
                            df_paracsv[column_name] = df_paracsv[column_name].astype(str).str.replace(' ', '').str.replace('i', '1', case=False).str.replace('l', '1', case=False)
                    
                    #TERMINAL VERIFICAÇÃO
                    column_name = None
                    normalized_columns = {}
                    for index in range(len(df_paracsv)):
                        for col in df_paracsv.columns:
                            normalized_value = str(df_paracsv.at[index, col]).lower().replace(' ', '').replace('i', '1').replace('l', '1')

                            if 'term1na1' == normalized_value:
                                column_name = col 
                                    
                            if column_name and column_name in df_paracsv.columns:
                                df_paracsv[column_name] = df_paracsv[column_name].astype(str).str.replace(' ', '').str.replace('i', '1', case=False).str.replace('l', '1', case=False)

                    for col in df_paracsv.columns:
                        normalized_col = str(col).lower().replace(' ', '').replace('i', '1').replace('l', '1')
                        normalized_columns[col] = normalized_col
                        
                        if normalized_col == 'term1na1':
                            column_name = col 

                        if column_name:
                            df_paracsv[column_name] = df_paracsv[column_name].astype(str).str.replace(' ', '').str.replace('i', '1', case=False).str.replace('l', '1', case=False)

                    #ADICIONANDO PLANILHAS A PLANILHA FINAL
                    file_path_csv = './arquivo.csv'
                    df_paracsv.to_csv(file_path_csv, index=False)

                    for sheet_name in wb_origem.sheetnames:
                        ws_origem = wb_origem[sheet_name]
                        ws_destino = wb_destino.create_sheet(title=table_name)

                        ws_destino.append(list(df_paracsv.columns))

                        for row in df_paracsv.itertuples(index=False):
                            ws_destino.append(row)

                    wb_source = openpyxl.load_workbook('./Cabos-Whirlpool.xlsx')
                    ws_source = wb_source.active

                    #VERIFICANDO CABOS POR COR E BITOLA
                    def mm2_to_awg(mm2):
                        conversion_table = {
                            '0,05 mm²': '30 AWG',
                            '0,08 mm²': '28 AWG',
                            '0,14 mm²': '26 AWG',
                            '0,20 mm²': '24 AWG',
                            '0,25 mm²': '24 AWG',
                            '0,32 mm²': '22 AWG',
                            '0,34 mm²': '22 AWG',
                            '0,35 mm²': '22 AWG',
                            '0,38 mm²': '21 AWG',
                            '0,50 mm²': '20 AWG',
                            '0,75 mm²': '18 AWG',
                            '1,00 mm²': '17 AWG',
                            '1,50 mm²': '16 AWG',
                            '2,50 mm²': '14 AWG',
                            '4,00 mm²': '12 AWG',
                            '6,00 mm²': '10 AWG',
                            '10,00 mm²': '8 AWG',
                            '16,00 mm²': '6 AWG',
                            '25,00 mm²': '4 AWG',
                            '35,00 mm²': '2 AWG',
                            '50,00 mm²': '1 AWG',
                        }
                        
                        return conversion_table.get(mm2, None)

                    def main_cables(mm2, color, temp):
                        find_cable = {
                            ('0,32 mm²', 'PR', '105') : '1022200802',
                            ('0,32 mm²', 'BR', '105') : '1022200402',
                            ('0,32 mm²', 'VM', '105') : '1022201101',
                            ('0,32 mm²', 'AZ', '105') : '1022200201',
                            ('0,32 mm²', 'LJ', '105') : '1022200600',
                            ('0,32 mm²', 'MR', '105') : '1022200701',
                            ('0,32 mm²', 'RO', '105') : '1022200900',
                            ('0,32 mm²', 'AM', '105') : '1022200101',
                            ('0,32 mm²', 'VD', '105') : '1022201001',
                            ('0,32 mm²', 'CZ', '105') : '1022200501',

                            ('0,20 mm²', 'MR', '105') : '1022400701',
                            ('0,20 mm²', 'VM', '105') : '1022401103',
                            ('0,20 mm²', 'PR', '105') : '1022400804',
                            ('0,20 mm²', 'RO', '105') : '1022400900',
                            ('0,20 mm²', 'AM', '105') : '1022400102',
                            ('0,20 mm²', 'AZ', '105') : '1022400201',
                            ('0,20 mm²', 'CZ', '105') : '1022400501',
                            ('0,20 mm²', 'LJ', '105') : '1022400601',
                            ('0,20 mm²', 'BR', '105') : '1022400401',
                            ('0,20 mm²', 'VD', '105') : '1022400100',


                            ('0,50 mm²', 'AZ', '105') : '1020500215',
                            ('0,50 mm²', 'VM', '105') : '1020501109',
                            ('0,50 mm²', 'MR', '105') : '1020500710',
                            ('0,50 mm²', 'VD/AM', '70') : '1010504400',
                            ('0,50 mm²', 'BR', '105') : '1020500410',
                            ('0,50 mm²', 'AZ', '70') : '125053',
                            ('0,50 mm²', 'BR', '70') : '125027',
                            ('0,50 mm²', 'VM', '70') : '125032',
                            ('0,50 mm²', 'PR', '70') : '125004',

                            ('0,75 mm²', 'VD/AM', '70') : '125076'

                        }
                        return find_cable.get((mm2, color, temp), None)
                    
                    for x in range(len(proc_data[0])):
                        cable = main_cables((proc_data[1][x]), (proc_data[0][x]), (proc_data[3][x]))
                        if cable:
                            proc_data[2].append(cable)
                        else:
                            for row in range(2, ws_source.max_row + 1):
                                temp_sheet =  ws_source.cell(row=row, column=3).value
                                cor_sheet = ws_source.cell(row=row, column=4).value  
                                bitola_sheet = ws_source.cell(row=row, column=5).value
                                item = ws_source.cell(row=row, column=1).value  
                                awg_value = mm2_to_awg(proc_data[1][x])
                                if proc_data[0][x] == cor_sheet and (proc_data[1][x] == bitola_sheet or awg_value == bitola_sheet) and (proc_data[3][x] + '°C') == temp_sheet:
                                    proc_data[2].append(item)
                                    break
                                else:
                                    item = None
                            if item == None:
                                proc_data[2].append('Item não encontrado!')
                                
                    if "Item" in wb_destino.sheetnames:
                        ws_item = wb_destino["Item"]
                        for cor, bitola, item in zip(proc_data[0], proc_data[1], proc_data[2]):
                            ws_item.append([cor, bitola, item])

                arquivo_excel_path = f'./Excel/planilha_final{self.filename_id}.xlsx'
                wb_destino.save(arquivo_excel_path)
                print('6', flush=True)

class Delete_Files():
    def __init__(self, contain_finalSheet: bool, path: str, filename_id: str):
        self.contain_finalSheet = contain_finalSheet
        self.path = path
        self.filename_id = filename_id
    
    def deleteFiles(self):
        if os.path.exists(self.path):
            for file in os.listdir(self.path):
                if self.filename_id in file:
                    if self.contain_finalSheet == True:
                        if 'planilha_final' not in file:
                            os.remove(os.path.join(self.path, file))
                    else:
                        os.remove(os.path.join(self.path, file))
    


                    

                    


                    



        



